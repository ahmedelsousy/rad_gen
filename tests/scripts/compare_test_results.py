"""Parse report.txt files under tests/data/<group>/outputs/<run>/ and emit
a side-by-side comparison as XLSX + Markdown.

Usage:
    python tests/scripts/compare_test_results.py \
        --xlsx results_comparison.xlsx \
        --md   results_comparison.md
"""

import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]

# (test_group, run_dir, friendly_label) — only completed full runs.
RUNS: List[Tuple[str, str, str]] = [
    ("single_wire", "l4_only_full",              "single_wire / L4 (Fs_mtx, N=1)"),
    ("single_wire", "l16_only_full",             "single_wire / L16 (Fs_mtx, N=1)"),
    ("stratix_iv",  "stratix_iv_rrg_debug",      "stratix_iv / RRG"),
    ("stratix_iv",  "stratix_iv_sb_muxes_debug", "stratix_iv / Method 1 (Explicit SB mux definitions)"),
    ("stratix_iv",  "stratix_iv_fs_mtx",   "stratix_iv / Method 2 (Fs Matrix)")
]


@dataclass
class MuxBlock:
    kind: str  # "SWITCH BLOCK" / "CONNECTION BLOCK" / "LOCAL MUX"
    style: str = ""
    required: str = ""
    implemented: str = ""
    lvl1: str = ""
    lvl2: str = ""
    unused: str = ""
    muxes_per_tile: str = ""
    sram_per_mux: str = ""


@dataclass
class Report:
    group: str
    run: str
    label: str
    path: Path

    created: str = ""
    run_options: Dict[str, str] = field(default_factory=dict)
    arch: Dict[str, str] = field(default_factory=dict)
    process: Dict[str, str] = field(default_factory=dict)
    sb_blocks: List[MuxBlock] = field(default_factory=list)
    cb_blocks: List[MuxBlock] = field(default_factory=list)
    local_blocks: List[MuxBlock] = field(default_factory=list)
    subcircuits: List[List[str]] = field(default_factory=list)  # rows
    subcircuit_header: List[str] = field(default_factory=list)
    tile_metal: Dict[str, str] = field(default_factory=dict)
    tile_area: List[Tuple[str, str, str]] = field(default_factory=list)  # block, area, frac
    vpr_delays: List[Tuple[str, str]] = field(default_factory=list)
    vpr_areas: List[Tuple[str, str]] = field(default_factory=list)
    summary: Dict[str, str] = field(default_factory=dict)
    hspice_sims: str = ""
    elapsed: str = ""


SECTION_RE = re.compile(r"^\s*-{3,}\s*$")


def _strip_comment_kv(line: str) -> Optional[Tuple[str, str]]:
    """Lines like '  vdd = 1.0' -> (vdd, 1.0)."""
    m = re.match(r"\s*([^=]+?)\s*=\s*(.+?)\s*$", line)
    if not m:
        return None
    return m.group(1).strip(), m.group(2).strip()


def _strip_colon_kv(line: str) -> Optional[Tuple[str, str]]:
    """Lines like '  Number of BLEs per cluster (N): 10' -> (key, value)."""
    m = re.match(r"\s*([^:]+?)\s*:\s*(.+?)\s*$", line)
    if not m:
        return None
    return m.group(1).strip(), m.group(2).strip()


def parse_report(group: str, run: str, label: str, path: Path) -> Report:
    rep = Report(group=group, run=run, label=label, path=path)
    text = path.read_text()
    lines = text.splitlines()

    i = 0
    while i < len(lines):
        line = lines[i]

        if line.startswith("Created "):
            rep.created = line[len("Created "):].strip()

        elif "RUN OPTIONS:" in line:
            i += 1
            while i < len(lines) and "ARCHITECTURE PARAMETERS:" not in lines[i]:
                kv = _strip_colon_kv(lines[i])
                if kv:
                    rep.run_options[kv[0]] = kv[1]
                i += 1
            continue

        elif "ARCHITECTURE PARAMETERS:" in line:
            i += 1
            while i < len(lines) and "PROCESS TECHNOLOGY PARAMETERS:" not in lines[i]:
                kv = _strip_colon_kv(lines[i])
                if kv:
                    rep.arch[kv[0]] = kv[1]
                i += 1
            continue

        elif "PROCESS TECHNOLOGY PARAMETERS:" in line:
            i += 1
            while i < len(lines) and "FPGA Implementation Details" not in lines[i]:
                kv = _strip_comment_kv(lines[i])
                if kv:
                    rep.process[kv[0]] = kv[1]
                i += 1
            continue

        elif "SWITCH BLOCK DETAILS:" in line:
            block, i = _parse_mux_block(lines, i + 1, "SWITCH BLOCK")
            rep.sb_blocks.append(block)
            continue

        elif "CONNECTION BLOCK DETAILS:" in line:
            block, i = _parse_mux_block(lines, i + 1, "CONNECTION BLOCK")
            rep.cb_blocks.append(block)
            continue

        elif "LOCAL MUX DETAILS:" in line:
            block, i = _parse_mux_block(lines, i + 1, "LOCAL MUX")
            rep.local_blocks.append(block)
            continue

        elif "SUBCIRCUIT AREA, DELAY & POWER" in line:
            i += 2  # skip header underline
            # next line: column header
            if i < len(lines):
                rep.subcircuit_header = re.split(r"\s{2,}", lines[i].strip())
                i += 1
            while i < len(lines):
                row_line = lines[i]
                if not row_line.strip():
                    break
                cols = re.split(r"\s{2,}", row_line.strip())
                if len(cols) >= 2:
                    rep.subcircuits.append(cols)
                i += 1
            continue

        elif line.strip().startswith("General routing metal pitch"):
            # Tile geometry block
            for j in range(i, min(i + 6, len(lines))):
                kv = _strip_comment_kv(lines[j])
                if kv:
                    rep.tile_metal[kv[0]] = kv[1]
                else:
                    txt = lines[j].strip()
                    if "Tile area" in txt:
                        rep.tile_metal["Tile area metal-limited"] = (
                            "NO" if "NOT" in txt else "YES"
                        )

        elif "TILE AREA CONTRIBUTIONS" in line:
            i += 3  # underline + col header
            while i < len(lines):
                row_line = lines[i]
                if not row_line.strip():
                    break
                cols = re.split(r"\s{2,}", row_line.strip())
                if len(cols) >= 3:
                    rep.tile_area.append((cols[0], cols[1], cols[2]))
                i += 1
            continue

        elif "VPR DELAYS" in line:
            i += 3  # underline + col header
            while i < len(lines):
                row_line = lines[i]
                if not row_line.strip():
                    break
                if "VPR AREAS" in row_line:
                    break
                # VPR delays are "  name   value"
                m = re.match(r"\s*(.+?)\s{2,}([0-9eE.+-]+)\s*$", row_line)
                if m:
                    rep.vpr_delays.append((m.group(1).strip(), m.group(2).strip()))
                i += 1
            continue

        elif "VPR AREAS" in line:
            i += 3
            while i < len(lines):
                row_line = lines[i]
                if not row_line.strip():
                    break
                if "SUMMARY" in row_line:
                    break
                m = re.match(r"\s*(.+?)\s{2,}([0-9eE.+-]+)\s*$", row_line)
                if m:
                    rep.vpr_areas.append((m.group(1).strip(), m.group(2).strip()))
                i += 1
            continue

        elif line.strip() == "SUMMARY":
            i += 2
            while i < len(lines):
                row_line = lines[i]
                if not row_line.strip() or row_line.startswith("|"):
                    break
                m = re.match(r"\s*(.+?)\s{2,}(.+?)\s*$", row_line)
                if m:
                    rep.summary[m.group(1).strip()] = m.group(2).strip()
                i += 1
            continue

        elif line.startswith("Number of HSPICE simulations performed:"):
            rep.hspice_sims = line.split(":", 1)[1].strip()

        elif line.startswith("Total time elapsed:"):
            rep.elapsed = line.split(":", 1)[1].strip()

        i += 1

    return rep


def _parse_mux_block(lines: List[str], start: int, kind: str) -> Tuple[MuxBlock, int]:
    block = MuxBlock(kind=kind)
    i = start
    while i < len(lines):
        text = lines[i].strip()
        if not text:
            i += 1
            break
        if any(
            text.startswith(t)
            for t in ("SWITCH BLOCK", "CONNECTION BLOCK", "LOCAL MUX",
                     "GEN ROUTING", "CLUSTER OUTPUT", "LUT DETAILS",
                     "LUT INPUT", "|", "Area and Delay")
        ):
            break
        kv = _strip_colon_kv(lines[i])
        if kv:
            k, v = kv
            kl = k.lower()
            if kl == "style":
                block.style = v
            elif "required mux size" in kl:
                block.required = v
            elif "implemented mux size" in kl:
                block.implemented = v
            elif "muxes per tile" in kl:
                block.muxes_per_tile = v
            elif "sram cells per mux" in kl:
                block.sram_per_mux = v
        else:
            m = re.match(r"\s*Level\s+1\s+size\s*=\s*(\S+)", lines[i])
            if m:
                block.lvl1 = m.group(1)
            m = re.match(r"\s*Level\s+2\s+size\s*=\s*(\S+)", lines[i])
            if m:
                block.lvl2 = m.group(1)
            m = re.match(r"\s*Number of unused inputs\s*=\s*(\S+)", lines[i])
            if m:
                block.unused = m.group(1)
        i += 1
    return block, i


# ---------- workbook construction ----------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True)
SUMMARY_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_FONT = Font(bold=True)


def _write_header(ws, row: int, headers: List[str]) -> int:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1


def _write_section(ws, row: int, title: str, span: int) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _autosize(ws, widths: List[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _collect_keys(dicts: List[Dict[str, str]]) -> List[str]:
    """Preserve first-seen ordering across dicts."""
    seen: Dict[str, None] = {}
    for d in dicts:
        for k in d.keys():
            seen.setdefault(k, None)
    return list(seen.keys())


def build_summary_sheet(wb: Workbook, reports: List[Report]) -> None:
    ws = wb.active
    ws.title = "Comparison"

    labels = [r.label for r in reports]
    headers = ["Field"] + labels
    n_cols = len(headers)

    row = 1
    row = _write_header(ws, row, headers)

    def write_kv(key: str, vals: List[str]):
        nonlocal row
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        for j, v in enumerate(vals, start=2):
            ws.cell(row=row, column=j, value=v)
        row += 1

    # Run identity
    row = _write_section(ws, row, "RUN IDENTITY", n_cols)
    write_kv("Test group",       [r.group for r in reports])
    write_kv("Run directory",    [r.run for r in reports])
    write_kv("Created",          [r.created for r in reports])
    write_kv("HSPICE sims",      [r.hspice_sims for r in reports])
    write_kv("Total elapsed",    [r.elapsed for r in reports])

    # Run options
    row = _write_section(ws, row, "RUN OPTIONS", n_cols)
    for k in _collect_keys([r.run_options for r in reports]):
        write_kv(k, [r.run_options.get(k, "") for r in reports])

    # Architecture
    row = _write_section(ws, row, "ARCHITECTURE PARAMETERS", n_cols)
    for k in _collect_keys([r.arch for r in reports]):
        write_kv(k, [r.arch.get(k, "") for r in reports])

    # Process tech
    row = _write_section(ws, row, "PROCESS TECHNOLOGY PARAMETERS", n_cols)
    for k in _collect_keys([r.process for r in reports]):
        write_kv(k, [r.process.get(k, "") for r in reports])

    # SB / CB / Local mux summary (per-block: keep aligned by index)
    row = _write_section(ws, row, "SWITCH BLOCK MUX(es)", n_cols)
    max_sb = max(len(r.sb_blocks) for r in reports)
    for idx in range(max_sb):
        for field in ("required", "implemented", "lvl1", "lvl2", "unused",
                      "muxes_per_tile", "sram_per_mux"):
            vals = []
            for r in reports:
                vals.append(getattr(r.sb_blocks[idx], field, "") if idx < len(r.sb_blocks) else "")
            write_kv(f"SB[{idx}] {field}", vals)

    row = _write_section(ws, row, "CONNECTION BLOCK MUX(es)", n_cols)
    max_cb = max(len(r.cb_blocks) for r in reports)
    for idx in range(max_cb):
        for field in ("required", "implemented", "lvl1", "lvl2", "unused",
                      "muxes_per_tile", "sram_per_mux"):
            vals = []
            for r in reports:
                vals.append(getattr(r.cb_blocks[idx], field, "") if idx < len(r.cb_blocks) else "")
            write_kv(f"CB[{idx}] {field}", vals)

    row = _write_section(ws, row, "LOCAL MUX", n_cols)
    max_lm = max(len(r.local_blocks) for r in reports)
    for idx in range(max_lm):
        for field in ("required", "implemented", "lvl1", "lvl2", "unused",
                      "muxes_per_tile", "sram_per_mux"):
            vals = []
            for r in reports:
                vals.append(getattr(r.local_blocks[idx], field, "") if idx < len(r.local_blocks) else "")
            write_kv(f"Local[{idx}] {field}", vals)

    # Tile geometry
    row = _write_section(ws, row, "TILE GEOMETRY", n_cols)
    for k in _collect_keys([r.tile_metal for r in reports]):
        write_kv(k, [r.tile_metal.get(k, "") for r in reports])

    # Tile area contributions
    row = _write_section(ws, row, "TILE AREA CONTRIBUTIONS (um^2 / frac)", n_cols)
    block_names: List[str] = []
    for r in reports:
        for b, _a, _f in r.tile_area:
            if b not in block_names:
                block_names.append(b)
    for bn in block_names:
        # one row for area, one for fraction
        area_vals = []
        frac_vals = []
        for r in reports:
            entry = next((x for x in r.tile_area if x[0] == bn), None)
            area_vals.append(entry[1] if entry else "")
            frac_vals.append(entry[2] if entry else "")
        write_kv(f"{bn} — area", area_vals)
        write_kv(f"{bn} — fraction", frac_vals)

    # VPR areas
    row = _write_section(ws, row, "VPR AREAS", n_cols)
    keys: List[str] = []
    for r in reports:
        for k, _v in r.vpr_areas:
            if k not in keys:
                keys.append(k)
    for k in keys:
        vals = []
        for r in reports:
            entry = next((x for x in r.vpr_areas if x[0] == k), None)
            vals.append(entry[1] if entry else "")
        write_kv(k, vals)

    # VPR delays
    row = _write_section(ws, row, "VPR DELAYS (s)", n_cols)
    keys = []
    for r in reports:
        for k, _v in r.vpr_delays:
            if k not in keys:
                keys.append(k)
    for k in keys:
        vals = []
        for r in reports:
            entry = next((x for x in r.vpr_delays if x[0] == k), None)
            vals.append(entry[1] if entry else "")
        write_kv(k, vals)

    # Summary (highlighted)
    summary_start = row
    row = _write_section(ws, row, "FINAL SUMMARY", n_cols)
    for k in _collect_keys([r.summary for r in reports]):
        cell = ws.cell(row=row, column=1, value=k)
        cell.font = SUMMARY_FONT
        cell.fill = SUMMARY_FILL
        for j, r in enumerate(reports, start=2):
            c = ws.cell(row=row, column=j, value=r.summary.get(k, ""))
            c.fill = SUMMARY_FILL
            c.font = SUMMARY_FONT
        row += 1

    # Freeze + column widths
    ws.freeze_panes = "B2"
    _autosize(ws, [42] + [28] * len(reports))


def build_subcircuit_sheet(wb: Workbook, reports: List[Report]) -> None:
    """One sheet per metric (Area / Delay / Power) with subcircuit rows x runs columns."""
    metric_cols = {
        "Area (um^2)": 1,   # index into row past Subcircuit name
        "Delay (ps)": 2,
        "tfall (ps)": 3,
        "trise (ps)": 4,
        "Power at 250MHz (uW)": 5,
    }
    for metric, idx in metric_cols.items():
        ws = wb.create_sheet(title=f"Subcircuits — {metric.split(' ')[0]}")
        headers = ["Subcircuit"] + [r.label for r in reports]
        row = _write_header(ws, 1, headers)

        # Collect all subcircuit names in order from first report that has them
        names: List[str] = []
        for r in reports:
            for cols in r.subcircuits:
                name = cols[0]
                if name not in names:
                    names.append(name)

        for name in names:
            ws.cell(row=row, column=1, value=name).font = Font(bold=True)
            for j, r in enumerate(reports, start=2):
                entry = next((c for c in r.subcircuits if c[0] == name), None)
                val = entry[idx] if entry and len(entry) > idx else ""
                ws.cell(row=row, column=j, value=val)
            row += 1

        ws.freeze_panes = "B2"
        _autosize(ws, [32] + [28] * len(reports))


# ---------- markdown ----------

def _md_table(headers: List[str], rows: List[List[str]]) -> str:
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def build_markdown(reports: List[Report]) -> str:
    labels = [r.label for r in reports]
    out: List[str] = []
    out.append("# Test Results Comparison\n")
    out.append("Source: `tests/data/<group>/outputs/<run>/report.txt` for each run.\n")
    out.append("Generated by `tests/scripts/compare_test_results.py`.\n")

    # Identity
    out.append("\n## Run identity\n")
    rows = [
        ["Test group"] + [r.group for r in reports],
        ["Run directory"] + [r.run for r in reports],
        ["Created"] + [r.created for r in reports],
        ["HSPICE sims"] + [r.hspice_sims for r in reports],
        ["Total elapsed"] + [r.elapsed for r in reports],
    ]
    out.append(_md_table(["Field"] + labels, rows))

    # Final summary (the headline)
    out.append("\n## Final summary (headline metrics)\n")
    keys = _collect_keys([r.summary for r in reports])
    rows = [[k] + [r.summary.get(k, "") for r in reports] for k in keys]
    out.append(_md_table(["Metric"] + labels, rows))

    # Architecture
    out.append("\n## Architecture parameters\n")
    keys = _collect_keys([r.arch for r in reports])
    rows = [[k] + [r.arch.get(k, "") for r in reports] for k in keys]
    out.append(_md_table(["Parameter"] + labels, rows))

    # SB blocks
    out.append("\n## Switch block mux(es)\n")
    max_sb = max(len(r.sb_blocks) for r in reports)
    rows = []
    for idx in range(max_sb):
        for f in ("required", "implemented", "lvl1", "lvl2", "unused", "muxes_per_tile", "sram_per_mux"):
            rows.append([f"SB[{idx}] {f}"] + [
                getattr(r.sb_blocks[idx], f, "") if idx < len(r.sb_blocks) else ""
                for r in reports
            ])
    out.append(_md_table(["Field"] + labels, rows))

    # CB blocks
    out.append("\n## Connection block mux(es)\n")
    max_cb = max(len(r.cb_blocks) for r in reports)
    rows = []
    for idx in range(max_cb):
        for f in ("required", "implemented", "lvl1", "lvl2", "unused", "muxes_per_tile", "sram_per_mux"):
            rows.append([f"CB[{idx}] {f}"] + [
                getattr(r.cb_blocks[idx], f, "") if idx < len(r.cb_blocks) else ""
                for r in reports
            ])
    out.append(_md_table(["Field"] + labels, rows))

    # Tile area
    out.append("\n## Tile area contributions (um^2)\n")
    block_names: List[str] = []
    for r in reports:
        for b, _a, _f in r.tile_area:
            if b not in block_names:
                block_names.append(b)
    rows = []
    for bn in block_names:
        row = [bn]
        for r in reports:
            entry = next((x for x in r.tile_area if x[0] == bn), None)
            row.append(f"{entry[1]} ({entry[2]})" if entry else "")
        rows.append(row)
    out.append(_md_table(["Block"] + labels, rows))

    # VPR
    out.append("\n## VPR areas\n")
    keys = []
    for r in reports:
        for k, _v in r.vpr_areas:
            if k not in keys:
                keys.append(k)
    rows = []
    for k in keys:
        row = [k]
        for r in reports:
            entry = next((x for x in r.vpr_areas if x[0] == k), None)
            row.append(entry[1] if entry else "")
        rows.append(row)
    out.append(_md_table(["VPR area"] + labels, rows))

    out.append("\n## VPR delays (s)\n")
    keys = []
    for r in reports:
        for k, _v in r.vpr_delays:
            if k not in keys:
                keys.append(k)
    rows = []
    for k in keys:
        row = [k]
        for r in reports:
            entry = next((x for x in r.vpr_delays if x[0] == k), None)
            row.append(entry[1] if entry else "")
        rows.append(row)
    out.append(_md_table(["Path"] + labels, rows))

    out.append("")
    return "\n".join(out)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(REPO_ROOT / "results_comparison.xlsx"))
    ap.add_argument("--md",   default=str(REPO_ROOT / "results_comparison.md"))
    args = ap.parse_args()

    reports: List[Report] = []
    for group, run, label in RUNS:
        path = REPO_ROOT / "tests" / "data" / group / "outputs" / run / "report.txt"
        if not path.exists():
            print(f"[skip] {path} not found")
            continue
        rep = parse_report(group, run, label, path)
        reports.append(rep)
        print(f"[ok]   parsed {group}/{run}: tile={rep.summary.get('Tile Area','?')}, "
              f"crit={rep.summary.get('Representative Critical Path Delay','?')}, "
              f"cost={rep.summary.get('Cost (area^1 x delay^1)') or rep.summary.get('Cost (area^1 x delay^2)','?')}")

    if not reports:
        print("No reports parsed.")
        return 1

    wb = Workbook()
    build_summary_sheet(wb, reports)
    build_subcircuit_sheet(wb, reports)
    wb.save(args.xlsx)
    print(f"\nXLSX written: {args.xlsx}")

    md = build_markdown(reports)
    Path(args.md).write_text(md)
    print(f"MD written:   {args.md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
